"""
core/report_engine.py
报表生成引擎骨架 — 基于 pandas / openpyxl，负责数据聚合与 Excel 导出。

后续在此模块填充具体业务逻辑，与 UI 层完全解耦。
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.field_config import build_formula_variables, load_field_tree_from_json

if TYPE_CHECKING:
    from core.database import DatabaseManager


class ReportEngine:
    """基于 pandas 的报表生成引擎。"""

    def __init__(self, db: "DatabaseManager"):
        self._db = db

    def generate_daily_df(self, start_date: str, end_date: str) -> pd.DataFrame:
        records = self._db.get_range(start_date, end_date)
        if not records:
            return pd.DataFrame(columns=["date", "total_profit"])
        df = pd.DataFrame(records)
        df["date"] = pd.to_datetime(df["date"])
        df = df[["date", "total_profit"]].sort_values("date")
        return df.reset_index(drop=True)

    def generate_monthly_df(self, year: int) -> pd.DataFrame:
        start = f"{year}-01-01"
        end = f"{year}-12-31"
        df = self.generate_daily_df(start, end)
        if df.empty:
            return pd.DataFrame(columns=["month", "total_profit"])
        df["month"] = df["date"].dt.month
        monthly = df.groupby("month", as_index=False)["total_profit"].sum()
        return monthly

    def generate_yearly_summary(self) -> pd.DataFrame:
        records = self._db.get_all_records()
        if not records:
            return pd.DataFrame(columns=["year", "total_profit"])
        df = pd.DataFrame(records)
        df["date"] = pd.to_datetime(df["date"])
        df["year"] = df["date"].dt.year
        yearly = df.groupby("year", as_index=False)["total_profit"].sum()
        return yearly

    def export_to_excel(self, export_path: str | Path, start_date: str, end_date: str) -> Path:
        export_path = Path(export_path)
        records = self._db.get_range(start_date, end_date)
        field_tree = load_field_tree_from_json(
            self._db.get_config(
                "custom_field_tree",
                self._db.get_config("custom_fields", '["营业额", "成本"]'),
            )
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "利润报表"

        self._write_report_header(sheet, start_date, end_date)

        if not records:
            sheet.cell(row=5, column=1, value="当前筛选区间没有数据。")
            workbook.save(export_path)
            return export_path

        current_row = 5
        grouped_records = self._group_records(records)

        for year, months in grouped_records:
            year_profit = self._sum_profit(
                record for _, month_records in months for record in month_records
            )
            sheet.cell(row=current_row, column=1, value=f"【年度标注】{year} 年")
            sheet.cell(row=current_row, column=7, value="年利润（元）")
            sheet.cell(row=current_row, column=8, value=round(year_profit, 2))
            self._style_section_row(sheet, current_row, kind="year")
            current_row += 1

            for month, month_records in months:
                month_profit = self._sum_profit(month_records)
                sheet.cell(row=current_row, column=1, value=f"【月份标注】{year} 年 {month:02d} 月")
                sheet.cell(row=current_row, column=7, value="月利润（元）")
                sheet.cell(row=current_row, column=8, value=round(month_profit, 2))
                self._style_section_row(sheet, current_row, kind="month")
                current_row += 1

                self._write_month_header(sheet, current_row)
                current_row += 1

                for record in month_records:
                    current_row = self._write_record_rows(sheet, current_row, record, field_tree)

                current_row += 1

            current_row += 2

        self._set_column_widths(sheet)
        workbook.save(export_path)
        return export_path

    def get_summary_data(self) -> dict[str, float]:
        import datetime

        now = datetime.date.today()
        today_str = now.strftime("%Y-%m-%d")

        today_rec = self._db.get_record(today_str)
        today_val = today_rec["total_profit"] if today_rec else 0.0

        month_start = now.replace(day=1).strftime("%Y-%m-%d")
        month_records = self._db.get_range(month_start, today_str)
        month_val = self._sum_profit(month_records)

        year_start = now.replace(month=1, day=1).strftime("%Y-%m-%d")
        year_records = self._db.get_range(year_start, today_str)
        year_val = self._sum_profit(year_records)

        return {
            "today": today_val,
            "month": month_val,
            "year": year_val,
        }

    def _group_records(self, records: list[dict]) -> list[tuple[int, list[tuple[int, list[dict]]]]]:
        grouped: dict[int, dict[int, list[dict]]] = {}
        for record in records:
            dt = pd.to_datetime(record["date"])
            grouped.setdefault(dt.year, {}).setdefault(dt.month, []).append(record)

        return [
            (year, [(month, grouped[year][month]) for month in sorted(grouped[year])])
            for year in sorted(grouped)
        ]

    def _write_report_header(self, sheet, start_date: str, end_date: str) -> None:
        sheet.merge_cells("A1:H1")
        sheet["A1"] = "利润分级报表"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        sheet.merge_cells("A2:H2")
        sheet["A2"] = (
            f"导出区间：{start_date} 至 {end_date}    "
            "标注说明：每月数据之间空 1 行，每年数据之间空 3 行，并展示月利润与年利润。"
        )
        sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A2"].font = Font(size=10, italic=True, color="666666")

    def _write_month_header(self, sheet, row: int) -> None:
        headers = ["日期", "一级分类", "二级分类", "金额（元）", "一级分类汇总（元）", "当日利润（元）", "标注", "备注"]
        for index, title in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=index, value=title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DCE6F1")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_record_rows(self, sheet, row: int, record: dict, field_tree: list[dict]) -> int:
        detail_rows = self._expand_record_details(record["raw_data"], field_tree)
        formula_totals = build_formula_variables(field_tree, record["raw_data"])

        if not detail_rows:
            detail_rows = [{"parent": "未分类", "child": "", "amount": 0.0}]

        start_row = row
        parent_ranges: list[tuple[int, int, str]] = []
        cursor = row
        idx = 0

        while idx < len(detail_rows):
            parent_name = str(detail_rows[idx]["parent"])
            block_start = cursor
            while idx < len(detail_rows) and str(detail_rows[idx]["parent"]) == parent_name:
                detail = detail_rows[idx]
                sheet.cell(row=cursor, column=3, value=detail["child"])
                sheet.cell(row=cursor, column=4, value=round(detail["amount"], 2))
                sheet.cell(row=cursor, column=7, value="二级分类明细" if detail["child"] else "一级分类明细")
                cursor += 1
                idx += 1

            block_end = cursor - 1
            parent_ranges.append((block_start, block_end, parent_name))

        end_row = cursor - 1

        sheet.cell(row=start_row, column=1, value=record["date"])
        sheet.cell(row=start_row, column=6, value=round(record["total_profit"], 2))
        sheet.cell(row=start_row, column=8, value=record.get("note", ""))
        if end_row > start_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            sheet.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
            sheet.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)

        for block_start, block_end, parent_name in parent_ranges:
            sheet.cell(row=block_start, column=2, value=parent_name)
            sheet.cell(row=block_start, column=5, value=round(formula_totals.get(parent_name, 0.0), 2))
            if block_end > block_start:
                sheet.merge_cells(start_row=block_start, start_column=2, end_row=block_end, end_column=2)
                sheet.merge_cells(start_row=block_start, start_column=5, end_row=block_end, end_column=5)

        for current_row in range(start_row, end_row + 1):
            for col in range(1, 9):
                sheet.cell(row=current_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        return end_row + 1

    def _expand_record_details(self, raw_data: dict, field_tree: list[dict]) -> list[dict[str, float | str]]:
        rows: list[dict[str, float | str]] = []
        seen_parents: set[str] = set()

        for item in field_tree:
            parent_name = item["name"]
            children = item["children"]
            seen_parents.add(parent_name)
            stored_value = raw_data.get(parent_name, {})

            if children:
                child_values = stored_value if isinstance(stored_value, dict) else {}
                for child_name in children:
                    rows.append({
                        "parent": parent_name,
                        "child": child_name,
                        "amount": float(child_values.get(child_name, 0.0) or 0.0),
                    })
            else:
                rows.append({
                    "parent": parent_name,
                    "child": "",
                    "amount": float(stored_value or 0.0),
                })

        for key, value in raw_data.items():
            if key in seen_parents:
                continue
            if isinstance(value, dict):
                for child_name, amount in value.items():
                    rows.append({
                        "parent": key,
                        "child": child_name,
                        "amount": float(amount or 0.0),
                    })
            else:
                rows.append({
                    "parent": key,
                    "child": "",
                    "amount": float(value or 0.0),
                })

        return rows

    def _style_section_row(self, sheet, row: int, kind: str) -> None:
        fill = "FCE4D6" if kind == "year" else "E2F0D9"
        for col in range(1, 9):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _sum_profit(self, records) -> float:
        return sum(float(record.get("total_profit", 0.0) or 0.0) for record in records)

    def _set_column_widths(self, sheet) -> None:
        widths = {
            1: 14,
            2: 18,
            3: 18,
            4: 14,
            5: 18,
            6: 14,
            7: 16,
            8: 18,
        }
        for column_index, width in widths.items():
            sheet.column_dimensions[get_column_letter(column_index)].width = width
